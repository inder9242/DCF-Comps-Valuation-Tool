# DCF model.py — simple final version

import io
from pathlib import Path
import pandas as pd
import yfinance as yf
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st

APP_TITLE = "DCF & Comps Valuation Tool"
EQUITY_FILE = "EQUITY_Final.xlsx"
TEMPLATE_FILE = "DCF_Template.xlsx"
MIN_PEERS = 10
MIN_POINTS = 20  # min weekly points to accept a price series

BASE = Path(__file__).resolve().parent

st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

left, right = st.columns([1.2, 0.8])
with left:
    symbol = st.text_input("An automated valuation tool for NSE-listed companies combining DCF, Trading Comps, and Football Field analysis with peer auto-selection, beta calculation, multi-method forecasting, and one-click Excel export. 
    Company Symbol (no .NS)", "TCS").strip().upper()
with right:
    run = st.button("Run Model")

# ---------- helpers ----------
def must_read_equity(path=BASE / EQUITY_FILE):
    if not path.exists():
        st.error(f"Missing {EQUITY_FILE} in this folder.")
        st.stop()
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip().str.lower()
    return df

def load_template(path=BASE / TEMPLATE_FILE):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    wb.active.title = "PEERS"
    for s in ["RAW_P_L", "RAW_B_S", "PRICE_HISTORY"]:
        wb.create_sheet(s)
    return wb

def unique_headers(headers):
    used, out = set(), []
    for i, h in enumerate(headers, 1):
        name = (str(h).strip() if h is not None else "")
        if not name:
            name = f"col_{i}"
        base, k = name, 1
        while name in used:
            k += 1
            name = f"{base}_{k}"
        used.add(name); out.append(name)
    return out

def robust_current_price(t):
    try:
        p = t.fast_info.get("lastPrice")
        if p: return float(p)
    except: pass
    try:
        p = t.info.get("currentPrice")
        if p: return float(p)
    except: pass
    try:
        h = t.history(period="5d")
        if not h.empty:
            return float(h["Close"].dropna().iloc[-1])
    except: pass
    return None

def pick_peers_min10(eq, base_row, target):
    # try strict -> relax: drop basicindustry -> industry -> sector -> macro -> all
    def filt(m=None, s=None, i=None, b=None):
        df = eq
        if m is not None: df = df[df["macro"] == m]
        if s is not None: df = df[df["sector"] == s]
        if i is not None: df = df[df["industry"] == i]
        if b is not None: df = df[df["basicindustry"] == b]
        lst = df["symbol"].dropna().astype(str).str.upper().unique().tolist()
        lst = [target] + [x for x in lst if x != target]  # target first
        return lst

    m, s, i, b = base_row["macro"], base_row["sector"], base_row["industry"], base_row["basicindustry"]
    for combo in [(m,s,i,b), (m,s,i,None), (m,s,None,None), (m,None,None,None), (None,None,None,None)]:
        peers = filt(*combo)
        if len(peers) >= MIN_PEERS:
            return peers
    return filt(None,None,None,None)

def grab_weekly_series(sym_full):
    # try 5y weekly -> fallback 1y daily resampled weekly
    df = yf.download(sym_full, period="5y", interval="1wk", progress=False, group_by="column")
    if isinstance(df.columns, pd.MultiIndex):
        use = "Adj Close" if "Adj Close" in set(df.columns.get_level_values(0)) else "Close"
        s = df[use]
        s = s.iloc[:,0] if isinstance(s, pd.DataFrame) else s
    else:
        if "Adj Close" in df.columns: s = df["Adj Close"]
        elif "Close" in df.columns:   s = df["Close"]
        else: s = pd.Series(dtype=float)
    s = s.dropna()
    if s.size >= MIN_POINTS:
        return s

    df = yf.download(sym_full, period="1y", interval="1d", progress=False, group_by="column")
    if "Adj Close" in df: s = df["Adj Close"].dropna()
    elif "Close" in df:   s = df["Close"].dropna()
    else:                 s = pd.Series(dtype=float)
    if not s.empty:
        s = s.to_frame("v")
        s.index = pd.to_datetime(s.index)
        s = s.resample("W-FRI").last()["v"].dropna()
        if s.size >= MIN_POINTS:
            return s
    return pd.Series(dtype=float)

def get_prices_first(peers, target):
    # build price table: Date + NIFTY50 + peers (only those with data)
    series = []
    nifty = grab_weekly_series("^NSEI")
    if not nifty.empty:
        nifty.name = "NIFTY50"; series.append(nifty)

    ok = []
    for p in peers:
        s = grab_weekly_series(p + ".NS")
        if s.empty: continue
        s.name = p
        series.append(s)
        ok.append(p)

    if not ok:
        return pd.DataFrame(), []

    df = series[0].to_frame()
    for s in series[1:]:
        df = df.join(s, how="outer")
    df = df.sort_index().reset_index().rename(columns={"index":"Date"})
    df["Date"] = pd.to_datetime(df["Date"]).dt.date

    if target in ok:
        ok = [target] + [x for x in ok if x != target]   # keep target first

    cols = ["Date"]
    if "NIFTY50" in df.columns: cols.append("NIFTY50")
    cols += ok
    return df[cols], ok

# ---- balance sheet alias getter ----
def bs_get(bal, latest, *aliases):
    if bal is None or bal.empty or latest is None:
        return None
    try:
        bal.index = bal.index.str.lower()
    except Exception:
        pass
    for a in aliases:
        a = a.lower()
        if a in bal.index:
            try:
                return bal.loc[a, latest]
            except Exception:
                pass
    return None

def fetch_snapshot(sym):
    try:
        t = yf.Ticker(sym + ".NS")
        info = t.info
        inc  = t.financials
        bal  = t.balance_sheet

        latest = inc.columns[0] if (inc is not None and not inc.empty) else None
        if inc is not None and not inc.empty: inc.index = inc.index.str.lower()
        if bal is not None and not bal.empty: bal.index = bal.index.str.lower()

        def g(df, key):
            try: return df.loc[key, latest] if (df is not None and latest in df.columns and key in df.index) else None
            except: return None

        # balance sheet items (aliases)
        cash_eq   = bs_get(bal, latest, "cash and cash equivalents", "cash & cash equivalents")
        restricted_cash = bs_get(bal, latest, "restricted cash")
        other_st_inv    = bs_get(bal, latest, "other short term investments", "short term investments", "other short-term investments")
        inv_fin_assets  = bs_get(bal, latest, "investment in financial assets", "investmentin financial assets", "investments in financial assets")
        afs_sec   = bs_get(bal, latest, "available for sale securities", "available-for-sale securities")
        other_invest    = bs_get(bal, latest, "other investments")
        goodwill        = bs_get(bal, latest, "goodwill")
        other_intang    = bs_get(bal, latest, "other intangible assets", "intangible assets other", "intangible assets")
        dta             = bs_get(bal, latest, "deferred tax assets", "non current deferred taxes assets")
        ncd_assets      = bs_get(bal, latest, "non-current deferred assets", "non current deferred assets", "deferred assets non current")
        fa_rr           = bs_get(bal, latest, "fixed assets revaluation reserve", "fixed asset revaluation reserve")
        stk_eq          = bs_get(bal, latest, "stockholders equity", "total stockholders equity", "shareholders equity", "total equity")
        minority        = bs_get(bal, latest, "minority interest", "noncontrolling interest", "non controlling interest")

        snap = {
            "Symbol": sym,
            "Market Cap": info.get("marketCap"),
            "No. of Shares": info.get("sharesOutstanding"),
            "Current Price": robust_current_price(t),
            "Total Revenue": g(inc, "total revenue"),
            "EBIT": g(inc, "ebit"),
            "EBITDA": g(inc, "ebitda"),
            "Cash and Cash Equivalents": cash_eq,
            "Restricted Cash": restricted_cash,
            "Other Short Term Investments": other_st_inv,
            "Investment in Financial Assets": inv_fin_assets,
            "Available For Sale Securities": afs_sec,
            "Other Investments": other_invest,
            "Goodwill": goodwill,
            "Other Intangible Assets": other_intang,
            "Deferred Tax Assets": dta,
            "Non-Current Deferred Assets": ncd_assets,
            "Fixed Asset Revaluation Reserve": fa_rr,
            "Stockholders Equity": stk_eq,
            "Minority Interest": minority,
            "Total Debt": info.get("totalDebt"),
            "Diluted EPS": info.get("trailingEps"),
        }
        return snap
    except Exception as e:
        st.warning(f"{sym}: {e}")
        return {"Symbol": sym}

def scrape_screener_tables(sym_no_ns):
    url = f"https://www.screener.in/company/{sym_no_ns}/consolidated/"
    try:
        r = requests.get(url, headers={"User-Agent":"Mozilla/5.0"}, timeout=20)
        if r.status_code != 200: return None, None
        soup = BeautifulSoup(r.text, "html.parser")
        def pick(title):
            h = soup.find("h2", string=lambda x: x and title in x.lower())
            return h.find_next("table", class_="data-table") if h else None
        def to_df(tbl):
            if not tbl: return None
            rows = tbl.find_all("tr")
            data = [[c.get_text(strip=True) for c in tr.find_all(["th","td"])] for tr in rows]
            if len(data) < 2: return None
            return pd.DataFrame(data[1:], columns=data[0])
        return to_df(pick("profit & loss")), to_df(pick("balance sheet"))
    except:
        return None, None

def write_keep_headers(ws, df, allow_new=False):
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if allow_new:
        for name in df.columns:
            if name not in headers:
                headers.append(name)
                ws.cell(row=1, column=len(headers), value=name)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    ordered = {h: (df[h] if h in df.columns else None) for h in headers}
    out = pd.DataFrame(ordered)
    for _, r in out.iterrows():
        ws.append(list(r.values))

def sort_beta_by_mcap(wb):
    sh = None
    for n in wb.sheetnames:
        if n.lower() == "beta":
            sh = wb[n]; break
    if sh is None: return
    rows = []
    maxc = sh.max_column
    for r in range(3, sh.max_row + 1):
        vals = [sh.cell(r, c).value for c in range(1, maxc+1)]
        m = sh.cell(r, 3).value
        try:
            m = float(str(m).replace(",", "")) if m is not None else float("-inf")
        except:
            m = float("-inf")
        rows.append((m, vals))
    rows.sort(key=lambda x: x[0], reverse=True)
    for i, (_, vals) in enumerate(rows, start=3):
        for c in range(1, maxc+1):
            sh.cell(i, c, vals[c-1])

# ---------- main ----------
if run:
    if not symbol:
        st.error("Enter a symbol."); st.stop()

    # 1) classification & peer pick
    eq = must_read_equity()
    row = eq[eq["symbol"].str.upper() == symbol]
    if row.empty:
        st.error(f"{symbol} not found in {EQUITY_FILE}."); st.stop()
    peers0 = pick_peers_min10(eq, row.iloc[0], symbol)

    # 2) price history FIRST; drop peers with no data (hard replace)
    with st.spinner("Downloading price history..."):
        price_df, peers = get_prices_first(peers0, symbol)
    if price_df.empty or len(peers) == 0:
        st.error("No peers have sufficient price history. Try another symbol."); st.stop()

    # 3) snapshots ONLY for peers with price history
    with st.spinner("Fetching peer snapshots..."):
        snaps = [fetch_snapshot(p) for p in peers]
        peers_df = pd.DataFrame(snaps)
        peers_df = peers_df.dropna(subset=["Total Debt", "No. of Shares"]).reset_index(drop=True)
        if peers_df.empty:
            st.error("Peers with price history failed fundamentals filter."); st.stop()

    # 4) Screener tables for target
    with st.spinner("Scraping financial statements..."):
        pl_df, bs_df = scrape_screener_tables(symbol)

    # 5) build workbook
    with st.spinner("Building workbook..."):
        wb = load_template()

        if "PEERS" in wb.sheetnames: write_keep_headers(wb["PEERS"], peers_df, allow_new=True)
        else:
            ws = wb.create_sheet("PEERS")
            for r in dataframe_to_rows(peers_df, index=False, header=True): ws.append(r)

        if pl_df is not None:
            if "RAW_P_L" in wb.sheetnames: write_keep_headers(wb["RAW_P_L"], pl_df, allow_new=False)
            else:
                ws = wb.create_sheet("RAW_P_L")
                for r in dataframe_to_rows(pl_df, index=False, header=True): ws.append(r)

        if bs_df is not None:
            if "RAW_B_S" in wb.sheetnames: write_keep_headers(wb["RAW_B_S"], bs_df, allow_new=False)
            else:
                ws = wb.create_sheet("RAW_B_S")
                for r in dataframe_to_rows(bs_df, index=False, header=True): ws.append(r)

        if "PRICE_HISTORY" in wb.sheetnames:
            ws = wb["PRICE_HISTORY"]; ws.delete_rows(1, ws.max_row)
            for r in dataframe_to_rows(price_df, index=False, header=True): ws.append(r)
        else:
            ws = wb.create_sheet("PRICE_HISTORY")
            for r in dataframe_to_rows(price_df, index=False, header=True): ws.append(r)

        sort_beta_by_mcap(wb)

        out_name = f"{symbol}_DCF_Comps_Valuation_Tool.xlsx"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    st.success(f"Workbook ready: {out_name}")

    # preview (simple, safe headers)
    st.subheader("Preview Workbook Sheets")
    prev = load_workbook(io.BytesIO(buf.getvalue()), data_only=True, read_only=True)
    tabs = st.tabs(prev.sheetnames)
    for t, name in zip(tabs, prev.sheetnames):
        with t:
            ws = prev[name]
            rows = list(ws.values)
            if not rows:
                st.write("(empty)"); continue
            cols = unique_headers([str(x) if x is not None else "" for x in rows[0]])
            df = pd.DataFrame(rows[1:], columns=cols).dropna(axis=1, how="all")
            st.dataframe(df, use_container_width=True)

    st.download_button("Download Excel", buf.getvalue(), file_name=out_name,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

